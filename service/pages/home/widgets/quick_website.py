import os

import flet as ft

from flet_core import ElevatedButton, Container, ResponsiveRow, Column, ListView, Row, Text, IconButton
from openpyxl import load_workbook
import pyperclip
from service.utils.common_utils import CommonUtils

from service.utils.file_utils import FileUtils
from service.utils.web_utils import WebUtils


class QuickWebsite(ft.UserControl):
    def __init__(self, parent: ft.Page):
        super().__init__()
        self.parent = parent
        # 获取配置文件地址
        self.configExcelPath = os.path.join(FileUtils.getAssetsPath(), "ShortcutWebsiteConfig.xlsx")

    def initData(self):
        self.shortcutWebsiteBtnList = []
        wb = load_workbook(filename=self.configExcelPath)

        ws = wb.active
        for cells in ws.iter_rows(min_row=2):  # 从第二行开始遍历
            print(cells[0].value, cells[1].value)
            btnItem = ElevatedButton(
                text=cells[0].value,
                col={"sm": 4},
                # 必须复制一份dirFilePath=row[1]
                on_click=lambda event, url=cells[1].value: self.openWebsite(event, url),
                on_long_press=lambda event, url=cells[1].value: self.handleLongPress(event, url),
            )
            self.shortcutWebsiteBtnList.append(btnItem)

    def build(self):
        self.initData()
        return ListView([
            Row([
                Text("快捷网址(长按复制网址)", size=20),
                IconButton(
                    icon=ft.icons.SETTINGS,
                    icon_color="blue400",
                    icon_size=20,
                    tooltip="快捷网址自定义",
                    on_click=lambda p: FileUtils.open_file_or_folder(self.configExcelPath)
                )
            ]),
            ResponsiveRow(self.shortcutWebsiteBtnList, alignment=ft.MainAxisAlignment.START),
        ])

    def openWebsite(self, event, url):
        print(event)
        WebUtils.open_url(url)
        pass

    def handleLongPress(self, event, url):
        print(event)
        pyperclip.copy(url)
        CommonUtils.showSnack(self.page, "网址已复制到剪切板")
        pass
