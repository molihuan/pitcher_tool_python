import os

import flet as ft
from flet_core import ElevatedButton, ResponsiveRow, Text, Row, ListView, IconButton
from openpyxl import load_workbook

from service.utils.common_utils import CommonUtils
from service.utils.file_utils import FileUtils


class QuickDirFile(ft.UserControl):
    def __init__(self, parent: ft.Page):
        super().__init__()
        self.parent = parent
        # 获取配置文件地址
        self.configExcelPath = os.path.join(FileUtils.getAssetsPath(), "ShortcutDirFileConfig.xlsx")

    def build(self):
        self.initData()
        return ListView([
            Row([
                Text("快捷文件/文件夹", size=20),
                IconButton(
                    icon=ft.icons.SETTINGS,
                    icon_color="blue400",
                    icon_size=20,
                    tooltip="快捷文件/文件夹自定义",
                    on_click=lambda p: FileUtils.open_file_or_folder(self.configExcelPath)
                )
            ]),
            ResponsiveRow(self.shortcutDirFileBtnList, alignment=ft.MainAxisAlignment.START),
        ])

    def initData(self):
        self.shortcutDirFileBtnList = []
        wb = load_workbook(filename=self.configExcelPath)
        ws = wb.active
        for cells in ws.iter_rows(min_row=2):  # 从第二行开始遍历
            print(cells[0].value, cells[1].value)
            btnItem = ElevatedButton(
                text=cells[0].value,
                col={"sm": 4},
                # 必须复制一份dirFilePath=row[1]
                on_click=lambda event, dirFilePath=cells[1].value: self.openDirFile(event, dirFilePath),
            )
            self.shortcutDirFileBtnList.append(btnItem)

    def openDirFile(self, event, dirFilePath):
        print(event)
        if not FileUtils.exists(dirFilePath):
            CommonUtils.showSnack(self.page, "路径不存在,请检查Excel配置")
            return
        FileUtils.open_file_or_folder(dirFilePath)
        pass
