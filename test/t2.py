from openpyxl import load_workbook

wb = load_workbook(filename='E:\\DesktopSpace\\Development\\Python\\pitcher_tool\\assets\\ShortcutDirFileConfig.xlsx')
ws = wb.active
for cells in ws.iter_rows(min_row=2):  # 从第二行开始遍历
    print(cells[0].value, cells[1].value)
    print()
